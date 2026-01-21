"""
Excel转Markdown脚本 v2.1
将Excel每个sheet转换为Markdown表格
支持合并单元格(使用HTML表格)、图片提取、隐藏Sheet标识、多Sheet分文件

使用方法：
    python convert_excel_to_md.py <输入Excel路径> <输出路径>

示例：
    python convert_excel_to_md.py "原始资料/数据.xlsx" "output_LLM/数据_v1.0.md"
    
功能特性：
    - 多Sheet支持（>10个自动分文件）
    - 合并单元格处理（HTML表格）
    - 图片提取（含悬浮图片、单元格内图片）
    - 隐藏Sheet标识
    - 空列自动过滤
"""
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
import sys
from pathlib import Path
import os
import re
from datetime import datetime

# Sheet数量阈值，超过此值则分文件输出
MULTI_FILE_THRESHOLD = 10

def sanitize_filename(name):
    """清理文件名中的非法字符"""
    return re.sub(r'[<>:"/\\|?*]', '_', name)

def get_merged_info(row, col, merged_ranges):
    """获取单元格的合并信息 (rowspan, colspan, is_merged_slave)"""
    for merged_range in merged_ranges:
        if (merged_range.min_row <= row <= merged_range.max_row and
            merged_range.min_col <= col <= merged_range.max_col):
            if row == merged_range.min_row and col == merged_range.min_col:
                rowspan = merged_range.max_row - merged_range.min_row + 1
                colspan = merged_range.max_col - merged_range.min_col + 1
                return rowspan, colspan, False
            else:
                return 1, 1, True
    return 1, 1, False

def get_image_at_position(images_map, row, col):
    """获取指定位置的图片"""
    key = (row, col)
    return images_map.get(key, None)

def extract_images_from_sheet(ws, sheet_name, output_dir):
    """从sheet中提取图片并保存，返回图片信息和位置映射"""
    images_info = []
    images_map = {}  # (row, col) -> image_info
    
    try:
        if not hasattr(ws, '_images') or not ws._images:
            return images_info, images_map
            
        for idx, image in enumerate(ws._images):
            # 获取图片位置
            anchor = image.anchor
            row, col = 1, 1
            if hasattr(anchor, '_from'):
                row = anchor._from.row + 1
                col = anchor._from.col + 1
            
            # 生成安全的文件名
            safe_sheet_name = sanitize_filename(sheet_name)
            img_filename = f"{safe_sheet_name}_image_{idx + 1}.png"
            img_path = output_dir / img_filename
            
            # 获取图片数据并保存
            try:
                if hasattr(image, '_data'):
                    img_data = image._data()
                    if isinstance(img_data, bytes):
                        with open(img_path, 'wb') as img_file:
                            img_file.write(img_data)
                    elif hasattr(img_data, 'read'):
                        with open(img_path, 'wb') as img_file:
                            img_file.write(img_data.read())
            except Exception as e:
                print(f"  保存图片失败: {e}")
                continue
            
            img_info = {
                'filename': img_filename,
                'path': str(img_path),
                'relative_path': f"images/{img_filename}",
                'row': row,
                'col': col
            }
            images_info.append(img_info)
            images_map[(row, col)] = img_info
            
    except Exception as e:
        print(f"  提取图片时出错: {e}")
    
    return images_info, images_map

def write_markdown_table(f, all_rows, non_empty_cols, header, images_map=None):
    """写入标准Markdown表格"""
    header_cells = [str(header[i]) if header[i] is not None else "" for i in non_empty_cols]
    f.write("| " + " | ".join(header_cells) + " |\n")
    f.write("|" + "|".join(["---" for _ in non_empty_cols]) + "|\n")
    
    for row_idx, row in enumerate(all_rows[1:], start=2):
        cells = []
        for col_idx in non_empty_cols:
            cell_value = str(row[col_idx]).replace("|", "\\|").replace("\n", " ") if row[col_idx] is not None else ""
            
            # 检查是否有嵌入图片
            if images_map:
                img_info = images_map.get((row_idx, col_idx + 1))
                if img_info:
                    if cell_value:
                        cell_value = f"{cell_value} ![图片]({img_info['relative_path']})"
                    else:
                        cell_value = f"![图片]({img_info['relative_path']})"
            
            cells.append(cell_value)
        f.write("| " + " | ".join(cells) + " |\n")

def write_html_table(f, ws, all_rows, non_empty_cols, merged_ranges, images_map=None):
    """写入HTML表格(支持合并单元格和嵌入图片)"""
    f.write("<table>\n")
    
    # 表头
    f.write("  <thead>\n    <tr>\n")
    header = all_rows[0]
    for col_idx in non_empty_cols:
        cell_value = str(header[col_idx]) if header[col_idx] is not None else ""
        f.write(f"      <th>{cell_value}</th>\n")
    f.write("    </tr>\n  </thead>\n")
    
    # 表体
    f.write("  <tbody>\n")
    for row_idx, row in enumerate(all_rows[1:], start=2):
        f.write("    <tr>\n")
        for col_idx in non_empty_cols:
            actual_col = col_idx + 1
            rowspan, colspan, is_slave = get_merged_info(row_idx, actual_col, merged_ranges)
            
            if is_slave:
                continue
            
            cell_value = str(row[col_idx]).replace("\n", "<br>") if row[col_idx] is not None else ""
            
            # 检查是否有嵌入图片
            if images_map:
                img_info = images_map.get((row_idx, actual_col))
                if img_info:
                    img_tag = f'<img src="{img_info["relative_path"]}" alt="图片">'
                    if cell_value:
                        cell_value = f"{cell_value}<br>{img_tag}"
                    else:
                        cell_value = img_tag
            
            attrs = ""
            if rowspan > 1:
                attrs += f' rowspan="{rowspan}"'
            if colspan > 1:
                attrs += f' colspan="{colspan}"'
            
            f.write(f"      <td{attrs}>{cell_value}</td>\n")
        f.write("    </tr>\n")
    f.write("  </tbody>\n</table>\n")

def process_sheet(ws, sheet_name, is_hidden, images_dir, images_map=None):
    """处理单个sheet，返回markdown内容"""
    content = []
    
    # 标题（含隐藏标识）
    hidden_mark = " [隐藏]" if is_hidden else ""
    content.append(f"## {sheet_name}{hidden_mark}  \n\n")
    
    # 提取图片
    images_info, sheet_images_map = extract_images_from_sheet(ws, sheet_name, images_dir)
    if images_map is not None:
        images_map.update(sheet_images_map)
    
    # 悬浮图片（不在表格单元格内的图片）
    if images_info:
        content.append("### 图片内容  \n\n")
        for img_info in images_info:
            content.append(f"![{img_info['filename']}]({img_info['relative_path']})  \n\n")
        content.append("> ⚠️ **注意**：上述图片需要根据 `图片识别.md` 的规则进行进一步识别处理。  \n\n")
    
    # 获取所有数据
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        content.append("(空表)  \n\n")
        return "".join(content), len(all_rows)
    
    # 检查哪些列全为空
    num_cols = len(all_rows[0]) if all_rows else 0
    non_empty_cols = []
    for col_idx in range(num_cols):
        col_has_data = any(row[col_idx] is not None and str(row[col_idx]).strip() for row in all_rows)
        if col_has_data:
            non_empty_cols.append(col_idx)
    
    if not non_empty_cols:
        content.append("(空表)  \n\n")
        return "".join(content), len(all_rows)
    
    # 获取合并单元格信息
    merged_ranges = list(ws.merged_cells.ranges)
    
    # 写入表格
    import io
    table_buffer = io.StringIO()
    
    if merged_ranges:
        write_html_table(table_buffer, ws, all_rows, non_empty_cols, merged_ranges, sheet_images_map)
    else:
        header = all_rows[0]
        write_markdown_table(table_buffer, all_rows, non_empty_cols, header, sheet_images_map)
    
    content.append(table_buffer.getvalue())
    content.append("\n")
    
    return "".join(content), len(all_rows)

def create_index_file(output_dir, excel_name, sheet_infos):
    """创建索引文件"""
    index_path = output_dir / "00_索引.md"
    
    with open(index_path, 'w', encoding='utf-8') as f:
        f.write(f"# {excel_name}  \n\n")
        f.write(f"> 来源文件: {excel_name}  \n")
        f.write(f"> 转换时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}  \n")
        f.write(f"> Sheet总数: {len(sheet_infos)}  \n\n")
        
        f.write("## Sheet列表  \n\n")
        f.write("| 序号 | Sheet名称 | 状态 | 行数 | 链接 |  \n")
        f.write("|------|-----------|------|------|------|  \n")
        
        for idx, info in enumerate(sheet_infos, 1):
            status = "**隐藏**" if info['is_hidden'] else "正常"
            f.write(f"| {idx} | {info['name']} | {status} | {info['rows']} | [查看]({info['filename']}) |  \n")
        
        f.write("\n---  \n\n")
        f.write("## 版本记录  \n\n")
        f.write("| 版本 | 日期 | 修改内容 | 修改人 |  \n")
        f.write("|------|------|----------|--------|  \n")
        f.write(f"| v1.0 | {datetime.now().strftime('%Y-%m-%d')} | 从Excel文件转换生成 | AI助手 |  \n")
    
    return index_path

def excel_to_markdown(excel_path, output_path):
    """将Excel文件转换为Markdown文件"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    excel_name = Path(excel_path).stem
    output_path = Path(output_path)
    
    total_sheets = len(wb.sheetnames)
    print(f"检测到 {total_sheets} 个Sheet")
    
    # 判断输出模式
    multi_file_mode = total_sheets > MULTI_FILE_THRESHOLD
    
    if multi_file_mode:
        # 多文件模式：创建输出目录
        output_dir = output_path.parent / f"{excel_name}_v1.0"
        output_dir.mkdir(exist_ok=True)
        images_dir = output_dir / "images"
        images_dir.mkdir(exist_ok=True)
        print(f"多文件模式: 输出到 {output_dir}")
    else:
        # 单文件模式
        output_dir = output_path.parent
        images_dir = output_dir / "images"
        images_dir.mkdir(exist_ok=True)
        print(f"单文件模式: 输出到 {output_path}")
    
    sheet_infos = []
    total_rows = 0
    
    for idx, sheet_name in enumerate(wb.sheetnames, 1):
        ws = wb[sheet_name]
        
        # 检查是否隐藏
        is_hidden = ws.sheet_state != 'visible'
        
        print(f"  处理 [{idx}/{total_sheets}] {sheet_name}{'(隐藏)' if is_hidden else ''}")
        
        # 处理sheet内容
        content, rows = process_sheet(ws, sheet_name, is_hidden, images_dir)
        total_rows += rows
        
        if multi_file_mode:
            # 多文件模式：每个sheet一个文件
            safe_name = sanitize_filename(sheet_name)
            hidden_suffix = "_隐藏" if is_hidden else ""
            filename = f"{idx:02d}_{safe_name}{hidden_suffix}.md"
            file_path = output_dir / filename
            
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(f"# {sheet_name}{'[隐藏]' if is_hidden else ''}  \n\n")
                f.write(f"> 来源: {excel_name} / {sheet_name}  \n\n")
                # 替换二级标题为三级标题
                f.write(content.replace("## ", "### ").replace("### 图片内容", "## 图片内容"))
            
            sheet_infos.append({
                'name': sheet_name,
                'filename': filename,
                'is_hidden': is_hidden,
                'rows': rows
            })
        else:
            # 单文件模式：追加到列表
            sheet_infos.append({
                'name': sheet_name,
                'content': content,
                'is_hidden': is_hidden,
                'rows': rows
            })
    
    if multi_file_mode:
        # 创建索引文件
        create_index_file(output_dir, excel_name, sheet_infos)
        print(f"\n转换完成: {output_dir}")
        print(f"共 {total_sheets} 个sheet, {total_rows} 行数据")
        return str(output_dir)
    else:
        # 写入单文件
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(f"# {excel_name}  \n\n")
            
            for info in sheet_infos:
                f.write(info['content'])
            
            f.write("---  \n\n")
            f.write("## 版本记录  \n\n")
            f.write("| 版本 | 日期 | 修改内容 | 修改人 |  \n")
            f.write("|------|------|----------|--------|  \n")
            f.write(f"| v1.0 | {datetime.now().strftime('%Y-%m-%d')} | 从Excel文件转换生成 | AI助手 |  \n")
        
        print(f"\n转换完成: {output_path}")
        print(f"共 {total_sheets} 个sheet, {total_rows} 行数据")
        return str(output_path)

if __name__ == "__main__":
    if len(sys.argv) >= 3:
        excel_path = sys.argv[1]
        output_path = sys.argv[2]
    elif len(sys.argv) == 2:
        excel_path = sys.argv[1]
        input_path = Path(excel_path)
        output_path = f"output_LLM/{input_path.stem}_v1.0.md"
    else:
        print("Excel转Markdown脚本 v2.1")
        print("=" * 50)
        print("用法: python convert_excel_to_md.py <输入Excel路径> [输出路径]")
        print()
        print("示例:")
        print('  python convert_excel_to_md.py "原始资料/数据.xlsx"')
        print('  python convert_excel_to_md.py "原始资料/数据.xlsx" "output_LLM/数据_v1.0.md"')
        print()
        print("功能特性:")
        print("  - ≤10个Sheet: 输出单个MD文件")
        print("  - >10个Sheet: 创建文件夹，每个Sheet一个MD文件")
        print("  - 支持合并单元格（HTML表格）")
        print("  - 支持图片提取")
        print("  - 隐藏Sheet自动标识")
        sys.exit(1)
    
    excel_to_markdown(excel_path, output_path)
