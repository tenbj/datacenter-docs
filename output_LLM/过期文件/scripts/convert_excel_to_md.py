"""
Excel转Markdown脚本 v2.0
将Excel每个sheet转换为Markdown表格
支持合并单元格(使用HTML表格)、图片提取
"""
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
import sys
from pathlib import Path
import os
import io

def get_merged_cell_value(ws, row, col, merged_ranges):
    """获取合并单元格的值"""
    for merged_range in merged_ranges:
        if (merged_range.min_row <= row <= merged_range.max_row and
            merged_range.min_col <= col <= merged_range.max_col):
            # 返回合并区域左上角的值
            return ws.cell(merged_range.min_row, merged_range.min_col).value
    return ws.cell(row, col).value

def get_merged_info(row, col, merged_ranges):
    """获取单元格的合并信息 (rowspan, colspan, is_merged_slave)"""
    for merged_range in merged_ranges:
        if (merged_range.min_row <= row <= merged_range.max_row and
            merged_range.min_col <= col <= merged_range.max_col):
            if row == merged_range.min_row and col == merged_range.min_col:
                # 这是合并区域的主单元格
                rowspan = merged_range.max_row - merged_range.min_row + 1
                colspan = merged_range.max_col - merged_range.min_col + 1
                return rowspan, colspan, False
            else:
                # 这是被合并的从属单元格
                return 1, 1, True
    return 1, 1, False

def has_merged_cells_in_sheet(ws):
    """检查sheet是否有合并单元格"""
    return len(ws.merged_cells.ranges) > 0

def extract_images_from_sheet(ws, sheet_name, output_dir):
    """从sheet中提取图片并保存"""
    images_info = []
    try:
        for idx, image in enumerate(ws._images):
            # 获取图片位置
            anchor = image.anchor
            if hasattr(anchor, '_from'):
                row = anchor._from.row + 1
                col = anchor._from.col + 1
            else:
                row = 1
                col = 1
            
            # 保存图片
            img_filename = f"{sheet_name}_image_{idx + 1}.png"
            img_path = output_dir / img_filename
            
            # 获取图片数据
            if hasattr(image, '_data'):
                img_data = image._data()
            elif hasattr(image, 'ref'):
                img_data = image.ref
            else:
                continue
            
            with open(img_path, 'wb') as img_file:
                if isinstance(img_data, bytes):
                    img_file.write(img_data)
                elif hasattr(img_data, 'read'):
                    img_file.write(img_data.read())
            
            images_info.append({
                'filename': img_filename,
                'path': str(img_path),
                'row': row,
                'col': col
            })
    except Exception as e:
        print(f"提取图片时出错: {e}")
    
    return images_info

def write_markdown_table(f, all_rows, non_empty_cols, header):
    """写入标准Markdown表格"""
    header_cells = [str(header[i]) if header[i] is not None else "" for i in non_empty_cols]
    f.write("| " + " | ".join(header_cells) + " |\n")
    f.write("|" + "|".join(["---" for _ in non_empty_cols]) + "|\n")
    
    for row in all_rows[1:]:
        cells = [str(row[i]).replace("|", "\\|").replace("\n", " ") if row[i] is not None else "" for i in non_empty_cols]
        f.write("| " + " | ".join(cells) + " |\n")

def write_html_table(f, ws, all_rows, non_empty_cols, merged_ranges):
    """写入HTML表格(支持合并单元格)"""
    f.write("<table>\n")
    
    # 表头
    f.write("  <thead>\n    <tr>\n")
    header = all_rows[0]
    for col_idx in non_empty_cols:
        cell_value = str(header[col_idx]) if header[col_idx] is not None else ""
        f.write(f"      <th>{cell_value}</th>\n")
    f.write("    </tr>\n  </thead>\n")
    
    # 表体
    f.write("  <tbody>\n")
    for row_idx, row in enumerate(all_rows[1:], start=2):
        f.write("    <tr>\n")
        for col_idx in non_empty_cols:
            actual_col = col_idx + 1  # openpyxl是1-indexed
            rowspan, colspan, is_slave = get_merged_info(row_idx, actual_col, merged_ranges)
            
            if is_slave:
                continue  # 跳过被合并的从属单元格
            
            cell_value = str(row[col_idx]).replace("\n", "<br>") if row[col_idx] is not None else ""
            
            attrs = ""
            if rowspan > 1:
                attrs += f' rowspan="{rowspan}"'
            if colspan > 1:
                attrs += f' colspan="{colspan}"'
            
            f.write(f"      <td{attrs}>{cell_value}</td>\n")
        f.write("    </tr>\n")
    f.write("  </tbody>\n</table>\n")

def excel_to_markdown(excel_path, output_path):
    """将Excel文件转换为Markdown文件"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    output_dir = Path(output_path).parent
    
    # 创建图片输出目录
    images_dir = output_dir / "images"
    images_dir.mkdir(exist_ok=True)
    
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(f"# {Path(excel_path).stem}  \n\n")
        
        total_sheets = len(wb.sheetnames)
        total_rows = 0
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            f.write(f"## {sheet_name}  \n\n")
            
            # 检查并提取图片
            if hasattr(ws, '_images') and ws._images:
                images_info = extract_images_from_sheet(ws, sheet_name, images_dir)
                if images_info:
                    f.write("### 图片内容  \n\n")
                    for img_info in images_info:
                        f.write(f"![{img_info['filename']}]({img_info['path']})  \n\n")
                    f.write("> ⚠️ **注意**：上述图片需要根据 `图片识别.md` 的规则进行进一步识别处理。  \n\n")
            
            # 获取所有数据
            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                f.write("(空表)  \n\n")
                continue
            
            total_rows += len(all_rows)
            
            # 检查哪些列全为空(特别检查首列)
            num_cols = len(all_rows[0]) if all_rows else 0
            non_empty_cols = []
            for col_idx in range(num_cols):
                col_has_data = any(row[col_idx] is not None and str(row[col_idx]).strip() for row in all_rows)
                if col_has_data:
                    non_empty_cols.append(col_idx)
            
            if not non_empty_cols:
                f.write("(空表)  \n\n")
                continue
            
            # 获取合并单元格信息
            merged_ranges = list(ws.merged_cells.ranges)
            
            # 根据是否有合并单元格选择输出格式
            if merged_ranges:
                # 使用HTML表格
                write_html_table(f, ws, all_rows, non_empty_cols, merged_ranges)
            else:
                # 使用标准Markdown表格
                header = all_rows[0]
                write_markdown_table(f, all_rows, non_empty_cols, header)
            
            f.write("\n")
        
        # 添加版本信息
        f.write("---  \n\n")
        f.write("## 版本记录  \n\n")
        f.write("| 版本 | 日期 | 修改内容 | 修改人 |  \n")
        f.write("|------|------|----------|--------|  \n")
        f.write("| v1.0 | 自动生成 | 从Excel文件转换生成 | AI助手 |  \n")
    
    print(f"转换完成: {output_path}")
    print(f"共 {total_sheets} 个sheet, {total_rows} 行数据")
    return output_path

if __name__ == "__main__":
    if len(sys.argv) >= 3:
        excel_path = sys.argv[1]
        output_path = sys.argv[2]
    else:
        # 默认测试路径
        excel_path = r'd:\GoogleAntigravityProjects\datacenter document\原始资料\其他\广告业务流程-创建.xlsm'
        output_path = r'd:\GoogleAntigravityProjects\datacenter document\output_LLM\广告业务流程-创建_v1.0.md'
    
    excel_to_markdown(excel_path, output_path)
