"""
Excel转Markdown脚本
将Excel每个sheet转换为Markdown表格
"""
import openpyxl
import sys
from pathlib import Path

def excel_to_markdown(excel_path, output_path):
    """将Excel文件转换为Markdown文件"""
    wb = openpyxl.load_workbook(excel_path)
    
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(f"# {Path(excel_path).stem}\n\n")
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            f.write(f"## {sheet_name}\n\n")
            
            # 获取所有数据
            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                f.write("(空表)\n\n")
                continue
            
            # 检查哪些列全为空
            num_cols = len(all_rows[0]) if all_rows else 0
            non_empty_cols = []
            for col_idx in range(num_cols):
                col_has_data = any(row[col_idx] is not None and str(row[col_idx]).strip() for row in all_rows)
                if col_has_data:
                    non_empty_cols.append(col_idx)
            
            if not non_empty_cols:
                f.write("(空表)\n\n")
                continue
            
            # 写入表头
            header = all_rows[0]
            header_cells = [str(header[i]) if header[i] is not None else "" for i in non_empty_cols]
            f.write("| " + " | ".join(header_cells) + " |\n")
            f.write("|" + "|".join(["---" for _ in non_empty_cols]) + "|\n")
            
            # 写入数据行
            for row in all_rows[1:]:
                cells = [str(row[i]).replace("|", "\\|").replace("\n", " ") if row[i] is not None else "" for i in non_empty_cols]
                f.write("| " + " | ".join(cells) + " |\n")
            
            f.write("\n")
    
    print(f"转换完成: {output_path}")
    print(f"共 {len(wb.sheetnames)} 个sheet, {len(all_rows)} 行数据")

if __name__ == "__main__":
    excel_path = r'd:\GoogleAntigravityProjects\datacenter document\个人文件\CK数据模型字段信息.xlsx'
    output_path = r'd:\GoogleAntigravityProjects\datacenter document\output_LLM\CK数据模型字段信息.md'
    excel_to_markdown(excel_path, output_path)
